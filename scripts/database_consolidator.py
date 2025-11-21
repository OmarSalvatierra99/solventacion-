"""
Consolidador de Base de Datos
Genera base de datos consolidada en Excel organizada por Ente y Fuente de Financiamiento
"""

import os
from typing import Dict, List, Optional
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re


class DatabaseConsolidator:
    """
    Consolida información de múltiples archivos en una base de datos Excel organizada
    Estructura: Ente > Fuente de Financiamiento > Histórico de Propuestas
    """

    def __init__(self):
        """Inicializa el consolidador de base de datos"""
        self.datos_consolidados = []

    def agregar_datos_archivo(self, metadatos: Dict, contenido_extraido: Dict):
        """
        Agrega los datos de un archivo procesado a la base de datos consolidada

        Args:
            metadatos: Metadatos extraídos del archivo
            contenido_extraido: Contenido completo extraído por los procesadores
        """
        ente = metadatos.get('ente', 'DESCONOCIDO')
        fuentes = metadatos.get('fuentes_financiamiento', ['NO_ESPECIFICADA'])
        periodo = metadatos.get('periodo', 'NO_ESPECIFICADO')
        tipo_doc = metadatos.get('tipo_documento', 'GENERAL')
        nombre_archivo = metadatos.get('nombre_archivo', '')

        # Extraer propuestas
        propuestas = contenido_extraido.get('contenido', {}).get('propuestas', [])

        # Si hay propuestas, agregar cada una
        if propuestas:
            for propuesta in propuestas:
                for fuente in fuentes:
                    registro = {
                        'Ente': ente,
                        'Fuente de Financiamiento': fuente,
                        'Periodo': periodo,
                        'Tipo Documento': tipo_doc,
                        'Archivo Origen': nombre_archivo,
                        'Número Propuesta': propuesta.get('numero'),
                        'Observación': self._limpiar_texto_para_excel(propuesta.get('observacion_texto', '')),
                        'Propuesta de Solventación': self._limpiar_texto_para_excel(propuesta.get('propuesta_texto', '')),
                        'Hoja' : propuesta.get('hoja', 'N/A'),
                        'Fila': propuesta.get('fila', 'N/A'),
                        'Fecha Procesamiento': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    }
                    self.datos_consolidados.append(registro)
        else:
            # Si no hay propuestas, agregar registro de archivo sin propuestas
            for fuente in fuentes:
                registro = {
                    'Ente': ente,
                    'Fuente de Financiamiento': fuente,
                    'Periodo': periodo,
                    'Tipo Documento': tipo_doc,
                    'Archivo Origen': nombre_archivo,
                    'Número Propuesta': 'N/A',
                    'Observación': 'Sin propuestas detectadas',
                    'Propuesta de Solventación': 'Sin propuestas detectadas en el archivo',
                    'Hoja': 'N/A',
                    'Fila': 'N/A',
                    'Fecha Procesamiento': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                self.datos_consolidados.append(registro)

    def _limpiar_texto_para_excel(self, texto: str) -> str:
        """
        Limpia y formatea texto para Excel

        Args:
            texto: Texto a limpiar

        Returns:
            Texto limpio y formateado
        """
        if not texto or texto == 'nan':
            return ''

        # Convertir a string si no lo es
        texto = str(texto)

        # Eliminar espacios múltiples
        texto = re.sub(r'\s+', ' ', texto)

        # Limitar longitud (Excel tiene límite de 32,767 caracteres por celda)
        if len(texto) > 32000:
            texto = texto[:32000] + '... [TRUNCADO]'

        return texto.strip()

    def generar_excel_consolidado(self, ruta_salida: str = 'base_datos_consolidada.xlsx'):
        """
        Genera el archivo Excel consolidado con múltiples hojas organizadas

        Args:
            ruta_salida: Ruta donde guardar el archivo Excel

        Returns:
            Ruta del archivo generado
        """
        if not self.datos_consolidados:
            raise ValueError("No hay datos para consolidar. Primero agregue datos con agregar_datos_archivo()")

        # Crear DataFrame
        df = pd.DataFrame(self.datos_consolidados)

        # Crear el libro de Excel
        wb = Workbook()
        wb.remove(wb.active)  # Eliminar hoja por defecto

        # Hoja 1: Todos los datos consolidados
        self._crear_hoja_principal(wb, df)

        # Hoja 2: Vista por Ente
        self._crear_hoja_por_ente(wb, df)

        # Hoja 3: Vista por Fuente de Financiamiento
        self._crear_hoja_por_fuente(wb, df)

        # Hoja 4: Resumen Estadístico
        self._crear_hoja_resumen(wb, df)

        # Guardar archivo
        wb.save(ruta_salida)
        print(f"✓ Base de datos consolidada generada: {ruta_salida}")
        return ruta_salida

    def _crear_hoja_principal(self, wb: Workbook, df: pd.DataFrame):
        """Crea la hoja principal con todos los datos"""
        ws = wb.create_sheet("Base de Datos Completa", 0)

        # Escribir encabezados
        headers = list(df.columns)
        ws.append(headers)

        # Estilo de encabezados
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Escribir datos
        for row in df.itertuples(index=False):
            ws.append(list(row))

        # Ajustar anchos de columna
        column_widths = {
            'A': 20,  # Ente
            'B': 25,  # Fuente
            'C': 15,  # Periodo
            'D': 20,  # Tipo Doc
            'E': 35,  # Archivo
            'F': 10,  # Número
            'G': 50,  # Observación
            'H': 60,  # Propuesta
            'I': 15,  # Hoja
            'J': 10,  # Fila
            'K': 20   # Fecha
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Aplicar bordes y alineación a todas las celdas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Congelar primera fila
        ws.freeze_panes = "A2"

    def _crear_hoja_por_ente(self, wb: Workbook, df: pd.DataFrame):
        """Crea hojas individuales por cada ente"""
        entes = df['Ente'].unique()

        for ente in sorted(entes):
            # Crear nombre de hoja válido (máximo 31 caracteres)
            nombre_hoja = f"Ente_{ente}"[:31]
            ws = wb.create_sheet(nombre_hoja)

            # Filtrar datos del ente
            df_ente = df[df['Ente'] == ente].copy()

            # Título de la hoja
            ws.merge_cells('A1:K1')
            titulo_cell = ws['A1']
            titulo_cell.value = f"ENTE: {ente}"
            titulo_cell.font = Font(bold=True, size=16, color="FFFFFF")
            titulo_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
            titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Escribir datos
            for r_idx, row in enumerate(dataframe_to_rows(df_ente, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Estilo de encabezado
                    if r_idx == 2:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Ajustar columnas
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                ws.column_dimensions[col].width = 20

            ws.column_dimensions['H'].width = 60  # Propuesta más ancha

    def _crear_hoja_por_fuente(self, wb: Workbook, df: pd.DataFrame):
        """Crea una hoja con resumen por fuente de financiamiento"""
        ws = wb.create_sheet("Por Fuente Financiamiento")

        # Título
        ws.merge_cells('A1:E1')
        titulo_cell = ws['A1']
        titulo_cell.value = "RESUMEN POR FUENTE DE FINANCIAMIENTO"
        titulo_cell.font = Font(bold=True, size=14, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Agrupar por fuente
        resumen = df.groupby(['Fuente de Financiamiento', 'Ente']).size().reset_index(name='Total Propuestas')

        # Escribir resumen
        for r_idx, row in enumerate(dataframe_to_rows(resumen, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

        # Ajustar columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20

    def _crear_hoja_resumen(self, wb: Workbook, df: pd.DataFrame):
        """Crea hoja con resumen estadístico"""
        ws = wb.create_sheet("Resumen Estadístico")

        # Título
        ws.merge_cells('A1:D1')
        titulo_cell = ws['A1']
        titulo_cell.value = "RESUMEN ESTADÍSTICO"
        titulo_cell.font = Font(bold=True, size=14, color="FFFFFF")
        titulo_cell.fill = PatternFill(start_color="203764", end_color="203764", fill_type="solid")
        titulo_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Estadísticas
        estadisticas = [
            ['Total de Archivos Procesados', df['Archivo Origen'].nunique()],
            ['Total de Entes', df['Ente'].nunique()],
            ['Total de Fuentes de Financiamiento', df['Fuente de Financiamiento'].nunique()],
            ['Total de Propuestas', len(df)],
            ['', ''],
            ['Distribución por Ente', ''],
        ]

        # Agregar distribución por ente
        entes_count = df['Ente'].value_counts()
        for ente, count in entes_count.items():
            estadisticas.append([f'  {ente}', count])

        estadisticas.append(['', ''])
        estadisticas.append(['Distribución por Fuente', ''])

        # Agregar distribución por fuente
        fuentes_count = df['Fuente de Financiamiento'].value_counts()
        for fuente, count in fuentes_count.items():
            estadisticas.append([f'  {fuente}', count])

        # Escribir estadísticas
        for idx, (label, value) in enumerate(estadisticas, start=3):
            ws[f'A{idx}'] = label
            ws[f'B{idx}'] = value

            # Estilo para títulos de sección
            if label and not label.startswith('  '):
                ws[f'A{idx}'].font = Font(bold=True, size=12)
                ws[f'A{idx}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

        # Ajustar columnas
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

    def limpiar_datos(self):
        """Limpia los datos consolidados para empezar una nueva consolidación"""
        self.datos_consolidados = []

    def obtener_estadisticas(self) -> Dict:
        """
        Obtiene estadísticas de los datos consolidados

        Returns:
            Diccionario con estadísticas
        """
        if not self.datos_consolidados:
            return {'error': 'No hay datos consolidados'}

        df = pd.DataFrame(self.datos_consolidados)

        return {
            'total_registros': len(df),
            'total_archivos': df['Archivo Origen'].nunique(),
            'total_entes': df['Ente'].nunique(),
            'total_fuentes': df['Fuente de Financiamiento'].nunique(),
            'entes': sorted(df['Ente'].unique().tolist()),
            'fuentes': sorted(df['Fuente de Financiamiento'].unique().tolist()),
            'distribucion_por_ente': df['Ente'].value_counts().to_dict(),
            'distribucion_por_fuente': df['Fuente de Financiamiento'].value_counts().to_dict()
        }


# Instancia global del consolidador
consolidator = DatabaseConsolidator()


def agregar_datos_archivo(metadatos: Dict, contenido_extraido: Dict):
    """
    Función de conveniencia para agregar datos de un archivo

    Args:
        metadatos: Metadatos extraídos del archivo
        contenido_extraido: Contenido completo extraído
    """
    consolidator.agregar_datos_archivo(metadatos, contenido_extraido)


def generar_excel_consolidado(ruta_salida: str = 'base_datos_consolidada.xlsx') -> str:
    """
    Función de conveniencia para generar el Excel consolidado

    Args:
        ruta_salida: Ruta donde guardar el archivo

    Returns:
        Ruta del archivo generado
    """
    return consolidator.generar_excel_consolidado(ruta_salida)


def obtener_estadisticas() -> Dict:
    """
    Función de conveniencia para obtener estadísticas

    Returns:
        Diccionario con estadísticas
    """
    return consolidator.obtener_estadisticas()


def limpiar_datos():
    """Función de conveniencia para limpiar datos"""
    consolidator.limpiar_datos()
