"""
Procesador de archivos XLSX
Extrae propuestas de solventación, metadatos, y detecta imágenes
"""

import os
from datetime import datetime
import pandas as pd
from html import escape
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage


def convertir_a_html_crudo(texto):
    """
    Convierte el texto de una celda en formato HTML crudo.
    El texto se sanitiza para HTML.

    Args:
        texto: Texto a convertir.

    Returns:
        str: Representación en HTML crudo del texto.
    """
    if not isinstance(texto, str) or not texto.strip():
        return ""  # Dejar en blanco si no hay contenido válido
    return f"<p style='text-align:justify'>{escape(texto)}</p>"


def detectar_imagenes_xlsx(filepath):
    """
    Detecta si el archivo XLSX contiene imágenes.

    Args:
        filepath (str): Ruta del archivo XLSX.

    Returns:
        dict: Información sobre imágenes en el archivo.
    """
    try:
        wb = load_workbook(filepath)
        imagenes = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Buscar imágenes en la hoja
            if hasattr(sheet, '_images'):
                for img in sheet._images:
                    imagenes.append({
                        'hoja': sheet_name,
                        'formato': img.format if hasattr(img, 'format') else 'desconocido'
                    })

        return {
            'tiene_imagenes': len(imagenes) > 0,
            'cantidad': len(imagenes),
            'detalles': imagenes
        }
    except Exception as e:
        return {
            'tiene_imagenes': False,
            'cantidad': 0,
            'detalles': [],
            'error_deteccion': str(e)
        }


def extraer_propuestas_xlsx(filepath):
    """
    Extrae propuestas de solventación de todas las hojas del archivo Excel.
    Ignora la primera aparición de "PROPUESTA DE SOLVENTACIÓN" y usa numeración global.

    Args:
        filepath (str): Ruta del archivo XLSX.

    Returns:
        list[dict]: Lista de propuestas extraídas con numeración global.
    """
    try:
        excel_data = pd.ExcelFile(filepath)
        datos_excel = []
        numero_global = 1

        for sheet_name in excel_data.sheet_names:
            df = excel_data.parse(sheet_name)
            primera_aparicion = True

            for i, row in df.iterrows():
                for idx in range(len(row) - 1):
                    cell_value = row.iloc[idx]

                    if isinstance(cell_value, str) and "PROPUESTA DE SOLVENTACIÓN" in cell_value:
                        # Ignorar la primera aparición
                        if primera_aparicion:
                            primera_aparicion = False
                            break

                        # Buscar observación (asumiendo que puede estar en columnas anteriores)
                        observacion = None
                        for obs_idx in range(max(0, idx - 2), idx):
                            obs_cell = row.iloc[obs_idx]
                            if isinstance(obs_cell, str) and "OBSERVACIÓN" in obs_cell:
                                if obs_idx + 1 < len(row):
                                    observacion = row.iloc[obs_idx + 1]
                                break

                        # Obtener la propuesta (celda a la derecha)
                        propuesta = row.iloc[idx + 1] if idx + 1 < len(row) else None
                        propuesta_html = convertir_a_html_crudo(propuesta)

                        datos_excel.append({
                            "numero": numero_global,
                            "hoja": sheet_name,
                            "observacion": convertir_a_html_crudo(observacion) if observacion else "Sin observación",
                            "propuesta_html": propuesta_html
                        })
                        numero_global += 1
                        break

        return datos_excel

    except Exception as e:
        raise Exception(f"Error al extraer propuestas: {str(e)}")


def extraer_metadatos_xlsx(filepath):
    """
    Extrae metadatos del archivo XLSX.

    Args:
        filepath (str): Ruta del archivo XLSX.

    Returns:
        dict: Metadatos del archivo.
    """
    try:
        wb = load_workbook(filepath)
        props = wb.properties

        # Obtener nombres de hojas
        nombres_hojas = wb.sheetnames
        total_hojas = len(nombres_hojas)

        # Detectar imágenes
        info_imagenes = detectar_imagenes_xlsx(filepath)

        return {
            'nombre_archivo': os.path.basename(filepath),
            'autor': props.creator or 'Desconocido',
            'titulo': props.title or 'Sin título',
            'asunto': props.subject or 'Sin asunto',
            'descripcion': props.description or 'Sin descripción',
            'fecha_creacion': props.created.isoformat() if props.created else None,
            'fecha_modificacion': props.modified.isoformat() if props.modified else None,
            'ultima_modificacion_por': props.lastModifiedBy or 'Desconocido',
            'total_hojas': total_hojas,
            'nombres_hojas': nombres_hojas,
            'tamano_archivo': os.path.getsize(filepath),
            'imagenes': info_imagenes
        }

    except Exception as e:
        return {
            'nombre_archivo': os.path.basename(filepath),
            'error_metadatos': str(e),
            'tamano_archivo': os.path.getsize(filepath)
        }


def calcular_estadisticas_xlsx(filepath, propuestas):
    """
    Calcula estadísticas del archivo XLSX.

    Args:
        filepath (str): Ruta del archivo XLSX.
        propuestas (list): Lista de propuestas extraídas.

    Returns:
        dict: Estadísticas del archivo.
    """
    try:
        excel_data = pd.ExcelFile(filepath)

        total_celdas_con_datos = 0
        total_formulas = 0
        total_filas = 0
        total_columnas = 0

        wb = load_workbook(filepath)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # Contar filas y columnas
            total_filas += sheet.max_row
            total_columnas += sheet.max_column

            # Contar celdas con datos y fórmulas
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        total_celdas_con_datos += 1

                        # Verificar si es una fórmula
                        if isinstance(cell.value, str) and cell.value.startswith('='):
                            total_formulas += 1

        return {
            'total_hojas': len(wb.sheetnames),
            'total_celdas_con_datos': total_celdas_con_datos,
            'total_formulas': total_formulas,
            'total_filas': total_filas,
            'total_columnas': total_columnas,
            'total_propuestas': len(propuestas)
        }

    except Exception as e:
        return {
            'error_estadisticas': str(e),
            'total_propuestas': len(propuestas)
        }


def process_xlsx(filepath):
    """
    Procesa un archivo XLSX y extrae toda la información relevante.

    Args:
        filepath (str): Ruta del archivo XLSX.

    Returns:
        dict: Información extraída del archivo.
    """
    try:
        # Extraer propuestas de solventación
        propuestas = extraer_propuestas_xlsx(filepath)

        # Extraer metadatos
        metadatos = extraer_metadatos_xlsx(filepath)

        # Calcular estadísticas
        estadisticas = calcular_estadisticas_xlsx(filepath, propuestas)

        return {
            'tipo_archivo': 'XLSX',
            'nombre_archivo': os.path.basename(filepath),
            'procesado_en': datetime.now().isoformat(),
            'metadatos': metadatos,
            'estadisticas': estadisticas,
            'contenido': {
                'propuestas': propuestas
            }
        }

    except Exception as e:
        return {
            'tipo_archivo': 'XLSX',
            'nombre_archivo': os.path.basename(filepath),
            'error': str(e),
            'procesado_en': datetime.now().isoformat()
        }
