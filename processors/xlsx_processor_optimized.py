"""
Procesador OPTIMIZADO de archivos XLSX
Extrae TODO el contenido fielmente: tablas completas, imágenes, formatos, estilos
Fallback a OpenAI solo cuando la extracción estructurada falla
"""

import os
import io
import base64
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
import pandas as pd
from html import escape
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Alignment
import re


class XLSXProcessorOptimized:
    """Procesador optimizado de archivos XLSX con extracción completa"""

    def __init__(self):
        self.use_openai_fallback = False
        self.openai_client = None

    def _init_openai(self):
        """Inicializa cliente OpenAI solo si es necesario"""
        if not self.openai_client:
            try:
                import os
                from openai import OpenAI
                from dotenv import load_dotenv
                load_dotenv()

                api_key = os.getenv('OPENAI_API_KEY')
                if api_key:
                    self.openai_client = OpenAI(api_key=api_key)
                    self.use_openai_fallback = True
            except Exception as e:
                print(f"OpenAI no disponible: {e}")
                self.use_openai_fallback = False

    def extraer_estilo_celda(self, cell) -> Dict[str, Any]:
        """Extrae información de estilo de una celda"""
        estilo = {}

        try:
            # Fuente
            if cell.font:
                # Extraer color de fuente correctamente
                color_fuente = None
                try:
                    if cell.font.color and hasattr(cell.font.color, 'rgb'):
                        color_value = cell.font.color.rgb
                        if isinstance(color_value, str):
                            color_fuente = color_value.strip()
                        else:
                            color_fuente = str(color_value)
                except Exception:
                    pass

                estilo['fuente'] = {
                    'nombre': cell.font.name,
                    'tamaño': cell.font.size,
                    'negrita': cell.font.bold,
                    'cursiva': cell.font.italic,
                    'subrayado': cell.font.underline is not None,
                    'color': color_fuente
                }

            # Relleno
            if cell.fill and cell.fill.start_color:
                try:
                    if hasattr(cell.fill.start_color, 'rgb'):
                        relleno_value = cell.fill.start_color.rgb
                        if isinstance(relleno_value, str):
                            estilo['relleno'] = relleno_value.strip()
                        else:
                            estilo['relleno'] = str(relleno_value)
                except Exception:
                    pass

            # Alineación
            if cell.alignment:
                estilo['alineacion'] = {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'ajustar_texto': cell.alignment.wrap_text
                }

            # Bordes
            if cell.border:
                estilo['bordes'] = True

            # Formato de número
            if cell.number_format:
                estilo['formato_numero'] = cell.number_format

        except Exception as e:
            # Silenciar errores de estilo
            pass

        return estilo

    def celda_a_html(self, cell, value) -> str:
        """Convierte una celda a HTML con estilos"""
        if value is None or value == '':
            return ''

        texto = escape(str(value))
        estilo = self.extraer_estilo_celda(cell)

        # Construir estilos CSS
        css_styles = ['text-align:left']

        if 'fuente' in estilo:
            fuente = estilo['fuente']
            if fuente.get('negrita'):
                texto = f"<b>{texto}</b>"
            if fuente.get('cursiva'):
                texto = f"<i>{texto}</i>"
            if fuente.get('subrayado'):
                texto = f"<u>{texto}</u>"
            if fuente.get('tamaño'):
                css_styles.append(f"font-size:{fuente['tamaño']}pt")
            if fuente.get('color'):
                # Formatear color correctamente
                color = fuente['color']
                if not color.startswith('#'):
                    # Remover los primeros 2 caracteres si son FF (alpha channel)
                    if len(color) > 6 and color[:2].upper() == 'FF':
                        color = color[2:]
                    color = f"#{color}"
                if len(color) == 7 or len(color) == 9:
                    css_styles.append(f"color:{color}")

        if 'relleno' in estilo and estilo['relleno']:
            # Formatear color de relleno correctamente
            relleno = estilo['relleno']
            if not relleno.startswith('#'):
                # Remover los primeros 2 caracteres si son FF (alpha channel)
                if len(relleno) > 6 and relleno[:2].upper() == 'FF':
                    relleno = relleno[2:]
                relleno = f"#{relleno}"
            if len(relleno) == 7 or len(relleno) == 9:
                css_styles.append(f"background-color:{relleno}")

        if 'alineacion' in estilo:
            alin = estilo['alineacion']
            if alin.get('horizontal'):
                css_styles.append(f"text-align:{alin['horizontal']}")
            if alin.get('ajustar_texto'):
                css_styles.append('white-space:pre-wrap')

        style_attr = '; '.join(css_styles)
        return f"<span style='{style_attr}'>{texto}</span>"

    def extraer_tabla_completa_hoja(self, sheet) -> str:
        """Extrae una hoja completa como tabla HTML con todos los estilos"""
        html = "<table border='1' style='border-collapse: collapse; width:100%;'>\n"

        # Procesar cada fila
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            html += "<tr>\n"

            for cell in row:
                # Obtener valor
                value = cell.value

                # Convertir a HTML con estilos
                contenido_html = self.celda_a_html(cell, value)

                # Verificar si es celda fusionada
                es_fusionada = False
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Solo mostrar contenido en la primera celda del rango fusionado
                        if cell.coordinate == merged_range.start_cell.coordinate:
                            rows_span = merged_range.max_row - merged_range.min_row + 1
                            cols_span = merged_range.max_col - merged_range.min_col + 1
                            rowspan_attr = f" rowspan='{rows_span}'" if rows_span > 1 else ""
                            colspan_attr = f" colspan='{cols_span}'" if cols_span > 1 else ""
                            html += f"<td{rowspan_attr}{colspan_attr}>{contenido_html}</td>\n"
                            es_fusionada = True
                        else:
                            # Celda oculta por fusión
                            es_fusionada = True
                        break

                if not es_fusionada:
                    html += f"<td>{contenido_html}</td>\n"

            html += "</tr>\n"

        html += "</table>"
        return html

    def extraer_imagenes_hoja(self, sheet) -> List[Dict[str, Any]]:
        """Extrae todas las imágenes de una hoja con sus datos binarios"""
        imagenes = []

        try:
            if hasattr(sheet, '_images'):
                for idx, img in enumerate(sheet._images):
                    # Obtener datos de la imagen
                    imagen_data = {
                        'indice': idx + 1,
                        'formato': img.format if hasattr(img, 'format') else 'desconocido',
                        'posicion': None,
                        'datos_base64': None
                    }

                    # Intentar obtener la posición
                    if hasattr(img, 'anchor') and hasattr(img.anchor, '_from'):
                        try:
                            imagen_data['posicion'] = {
                                'columna': img.anchor._from.col,
                                'fila': img.anchor._from.row
                            }
                        except:
                            pass

                    # Intentar obtener los datos binarios de la imagen
                    try:
                        if hasattr(img, '_data'):
                            imagen_bytes = img._data()
                            imagen_data['datos_base64'] = base64.b64encode(imagen_bytes).decode('utf-8')
                            imagen_data['tamaño_bytes'] = len(imagen_bytes)
                    except Exception as e:
                        imagen_data['error_extraccion'] = str(e)

                    imagenes.append(imagen_data)

        except Exception as e:
            print(f"Error al extraer imágenes: {e}")

        return imagenes

    def extraer_informacion_adicional(self, texto: str) -> Dict[str, Any]:
        """
        Extrae información adicional del texto de propuestas y observaciones
        """
        info = {
            'fechas': [],
            'responsables': [],
            'referencias': [],
            'palabras_clave': []
        }

        # Buscar fechas (varios formatos)
        patrones_fecha = [
            r'\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\b',
            r'\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b',
            r'\b(?:enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:de\s+)?\d{4}\b',
            r'\b(?:ene|feb|mar|abr|may|jun|jul|ago|sep|oct|nov|dic)[a-z]*\.?\s+\d{4}\b'
        ]

        for patron in patrones_fecha:
            fechas_encontradas = re.findall(patron, texto, re.IGNORECASE)
            info['fechas'].extend(fechas_encontradas)

        # Buscar responsables
        palabras_responsable = ['responsable:', 'encargado:', 'titular:', 'director:', 'coordinador:', 'jefe:']
        for palabra in palabras_responsable:
            if palabra in texto.lower():
                inicio = texto.lower().find(palabra)
                fragmento = texto[inicio:inicio+100]
                info['responsables'].append(fragmento)

        # Buscar referencias numéricas
        referencias = re.findall(r'\b(?:ref|referencia|no|número|num)\.?\s*:?\s*(\d+(?:[/-]\d+)*)\b', texto, re.IGNORECASE)
        info['referencias'].extend(referencias)

        # Palabras clave importantes
        palabras_importantes = ['cumplimiento', 'incumplimiento', 'pendiente', 'realizado', 'en proceso',
                                'evidencia', 'documentación', 'plazo', 'vencimiento', 'urgente', 'prioritario']
        for palabra in palabras_importantes:
            if palabra in texto.lower():
                info['palabras_clave'].append(palabra)

        return info

    def extraer_propuestas_estructuradas(self, filepath: str) -> List[Dict[str, Any]]:
        """
        Extrae propuestas usando lógica estructurada (método mejorado con información adicional)
        """
        try:
            excel_data = pd.ExcelFile(filepath)
            wb = load_workbook(filepath)
            propuestas = []
            numero_global = 1

            for sheet_name in excel_data.sheet_names:
                df = excel_data.parse(sheet_name)
                sheet = wb[sheet_name]

                # Buscar todas las apariciones de "PROPUESTA DE SOLVENTACIÓN"
                for i, row in df.iterrows():
                    for idx in range(len(row) - 1):
                        cell_value = row.iloc[idx]

                        if isinstance(cell_value, str) and "PROPUESTA DE SOLVENTACIÓN" in cell_value.upper():
                            # Buscar número de referencia o clasificación al inicio
                            numero_referencia = None
                            clasificacion = None
                            if idx > 0:
                                primera_celda = row.iloc[0]
                                if primera_celda and str(primera_celda).strip():
                                    primer_valor = str(primera_celda).strip()
                                    if re.match(r'^\d+(\.\d+)*$', primer_valor):
                                        numero_referencia = primer_valor
                                    elif re.match(r'^[A-Z\d\-_/]+$', primer_valor):
                                        clasificacion = primer_valor

                            # Buscar observación
                            observacion = None
                            observacion_html = None

                            for obs_idx in range(max(0, idx - 3), idx + 1):
                                obs_cell = row.iloc[obs_idx] if obs_idx < len(row) else None
                                if obs_cell and isinstance(obs_cell, str) and "OBSERVACIÓN" in obs_cell.upper():
                                    if obs_idx + 1 < len(row):
                                        obs_value = row.iloc[obs_idx + 1]
                                        observacion = self._limpiar_texto(obs_value)

                                        # Obtener celda original para HTML con estilo
                                        try:
                                            cell_obs = sheet.cell(row=i+2, column=obs_idx+2)
                                            observacion_html = self.celda_a_html(cell_obs, obs_value)
                                        except:
                                            observacion_html = f"<p>{escape(str(obs_value))}</p>"
                                    break

                            # Obtener propuesta
                            propuesta = None
                            propuesta_html = None

                            if idx + 1 < len(row):
                                prop_value = row.iloc[idx + 1]
                                propuesta = self._limpiar_texto(prop_value)

                                # Obtener celda original para HTML con estilo
                                try:
                                    cell_prop = sheet.cell(row=i+2, column=idx+2)
                                    propuesta_html = self.celda_a_html(cell_prop, prop_value)
                                except:
                                    propuesta_html = f"<p>{escape(str(prop_value))}</p>"

                            if propuesta and propuesta.strip():
                                # Extraer información adicional
                                texto_completo = f"{observacion or ''} {propuesta}"
                                info_adicional = self.extraer_informacion_adicional(texto_completo)

                                propuesta_data = {
                                    "numero": numero_global,
                                    "hoja": sheet_name,
                                    "fila": i + 2,
                                    "observacion_texto": observacion or "Sin observación",
                                    "observacion_html": observacion_html or "<p>Sin observación</p>",
                                    "propuesta_texto": propuesta,
                                    "propuesta_html": propuesta_html,
                                    "metodo_extraccion": "estructurado"
                                }

                                # Agregar información adicional si existe
                                if numero_referencia:
                                    propuesta_data['referencia'] = numero_referencia
                                if clasificacion:
                                    propuesta_data['clasificacion'] = clasificacion
                                if info_adicional['fechas']:
                                    propuesta_data['fechas_encontradas'] = info_adicional['fechas']
                                if info_adicional['responsables']:
                                    propuesta_data['responsables_mencionados'] = info_adicional['responsables']
                                if info_adicional['referencias']:
                                    propuesta_data['referencias_numericas'] = info_adicional['referencias']
                                if info_adicional['palabras_clave']:
                                    propuesta_data['palabras_clave'] = info_adicional['palabras_clave']

                                propuestas.append(propuesta_data)
                                numero_global += 1
                            break

            return propuestas

        except Exception as e:
            raise Exception(f"Error en extracción estructurada: {str(e)}")

    def extraer_con_openai(self, filepath: str, hoja_html: str, sheet_name: str) -> List[Dict[str, Any]]:
        """
        Fallback: usa OpenAI para extraer propuestas cuando la lógica estructurada falla
        """
        if not self.use_openai_fallback:
            self._init_openai()

        if not self.openai_client:
            return []

        try:
            prompt = f"""Eres un experto en análisis de documentos de auditoría y solventación.

Analiza la siguiente tabla de Excel (en formato HTML) y extrae TODAS las propuestas de solventación que encuentres.

Una propuesta típicamente tiene:
1. Una OBSERVACIÓN (opcional)
2. Una PROPUESTA DE SOLVENTACIÓN

TABLA DE LA HOJA "{sheet_name}":
{hoja_html[:8000]}  # Limitar a 8000 caracteres para no exceder límites

Extrae TODAS las propuestas y devuelve un JSON con este formato:
{{
    "propuestas": [
        {{
            "numero": 1,
            "observacion": "texto de la observación o 'Sin observación'",
            "propuesta": "texto de la propuesta"
        }},
        ...
    ]
}}

IMPORTANTE: Solo devuelve el JSON, sin texto adicional."""

            response = self.openai_client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Eres un experto en análisis de documentos. Respondes solo en JSON válido."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,
                max_tokens=2000
            )

            # Parsear respuesta
            import json
            respuesta_texto = response.choices[0].message.content.strip()

            # Limpiar markdown si existe
            if respuesta_texto.startswith('```'):
                respuesta_texto = respuesta_texto.split('```')[1]
                if respuesta_texto.startswith('json'):
                    respuesta_texto = respuesta_texto[4:]

            resultado = json.loads(respuesta_texto)

            # Formatear propuestas
            propuestas = []
            for idx, prop in enumerate(resultado.get('propuestas', []), start=1):
                propuestas.append({
                    "numero": idx,
                    "hoja": sheet_name,
                    "observacion_texto": prop.get('observacion', 'Sin observación'),
                    "observacion_html": f"<p>{escape(prop.get('observacion', 'Sin observación'))}</p>",
                    "propuesta_texto": prop.get('propuesta', ''),
                    "propuesta_html": f"<p>{escape(prop.get('propuesta', ''))}</p>",
                    "metodo_extraccion": "openai_fallback"
                })

            return propuestas

        except Exception as e:
            print(f"Error en extracción con OpenAI: {e}")
            return []

    def _limpiar_texto(self, texto: Any) -> str:
        """Limpia y normaliza texto extraído"""
        if texto is None:
            return ""
        if not isinstance(texto, str):
            texto = str(texto)
        # Eliminar espacios múltiples y saltos de línea excesivos
        texto = re.sub(r'\s+', ' ', texto).strip()
        return texto

    def extraer_metadatos(self, filepath: str) -> Dict[str, Any]:
        """Extrae metadatos completos del archivo XLSX"""
        try:
            wb = load_workbook(filepath)
            props = wb.properties

            # Obtener información de hojas
            hojas_info = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                hojas_info.append({
                    'nombre': sheet_name,
                    'filas': sheet.max_row,
                    'columnas': sheet.max_column,
                    'celdas_activas': sum(1 for row in sheet.iter_rows() for cell in row if cell.value is not None)
                })

            return {
                'nombre_archivo': os.path.basename(filepath),
                'autor': props.creator or 'Desconocido',
                'titulo': props.title or 'Sin título',
                'asunto': props.subject or 'Sin asunto',
                'descripcion': props.description or 'Sin descripción',
                'fecha_creacion': props.created.isoformat() if props.created else None,
                'fecha_modificacion': props.modified.isoformat() if props.modified else None,
                'ultima_modificacion_por': props.lastModifiedBy or 'Desconocido',
                'total_hojas': len(wb.sheetnames),
                'nombres_hojas': wb.sheetnames,
                'hojas_info': hojas_info,
                'tamano_archivo': os.path.getsize(filepath)
            }

        except Exception as e:
            return {
                'nombre_archivo': os.path.basename(filepath),
                'error_metadatos': str(e),
                'tamano_archivo': os.path.getsize(filepath) if os.path.exists(filepath) else 0
            }

    def process_xlsx(self, filepath: str) -> Dict[str, Any]:
        """
        Procesa un archivo XLSX de manera optimizada
        Extrae TODO el contenido fielmente y usa OpenAI solo como fallback
        """
        try:
            wb = load_workbook(filepath)

            # 1. Extraer metadatos
            metadatos = self.extraer_metadatos(filepath)

            # 2. Extraer contenido completo de cada hoja
            hojas_completas = []
            todas_propuestas = []
            total_imagenes = 0

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]

                # Extraer tabla completa con estilos
                tabla_html = self.extraer_tabla_completa_hoja(sheet)

                # Extraer imágenes de la hoja
                imagenes = self.extraer_imagenes_hoja(sheet)
                total_imagenes += len(imagenes)

                hojas_completas.append({
                    'nombre': sheet_name,
                    'contenido_html': tabla_html,
                    'imagenes': imagenes,
                    'total_filas': sheet.max_row,
                    'total_columnas': sheet.max_column
                })

            # 3. Intentar extracción estructurada de propuestas
            propuestas = []
            try:
                propuestas = self.extraer_propuestas_estructuradas(filepath)
            except Exception as e:
                print(f"Extracción estructurada falló: {e}")

            # 4. Si no se encontraron propuestas, intentar con OpenAI (fallback)
            if len(propuestas) == 0 and self.use_openai_fallback:
                print("Usando OpenAI como fallback...")
                for hoja in hojas_completas:
                    propuestas_openai = self.extraer_con_openai(
                        filepath,
                        hoja['contenido_html'],
                        hoja['nombre']
                    )
                    propuestas.extend(propuestas_openai)

            # 5. Calcular estadísticas
            estadisticas = {
                'total_hojas': len(wb.sheetnames),
                'total_propuestas': len(propuestas),
                'total_imagenes': total_imagenes,
                'total_filas': sum(h['total_filas'] for h in hojas_completas),
                'total_columnas': sum(h['total_columnas'] for h in hojas_completas),
                'metodo_extraccion_usado': 'estructurado' if any(p.get('metodo_extraccion') == 'estructurado' for p in propuestas) else 'openai_fallback' if propuestas else 'sin_propuestas'
            }

            return {
                'tipo_archivo': 'XLSX',
                'nombre_archivo': os.path.basename(filepath),
                'procesado_en': datetime.now().isoformat(),
                'metadatos': metadatos,
                'estadisticas': estadisticas,
                'contenido': {
                    'hojas_completas': hojas_completas,
                    'propuestas': propuestas
                },
                'extraccion_exitosa': True
            }

        except Exception as e:
            return {
                'tipo_archivo': 'XLSX',
                'nombre_archivo': os.path.basename(filepath),
                'error': str(e),
                'procesado_en': datetime.now().isoformat(),
                'extraccion_exitosa': False
            }


# Instancia global del procesador
processor = XLSXProcessorOptimized()


def process_xlsx(filepath: str) -> Dict[str, Any]:
    """Función de compatibilidad con el código existente"""
    return processor.process_xlsx(filepath)
